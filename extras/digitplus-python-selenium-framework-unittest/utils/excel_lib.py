import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from multipledispatch import dispatch

class ExcelLib:
    @dispatch(str)
    def get_row_count(self, file_path):
        sheet = openpyxl.load_workbook(file_path).active
        return sheet.max_row

    @dispatch(str, str)
    def get_row_count(self, file_path, sheet_name):
        sheet = openpyxl.load_workbook(file_path).get_sheet_by_name(sheet_name)
        return sheet.max_row

    @dispatch(str)
    def get_col_count(self, file_path):
        sheet = openpyxl.load_workbook(file_path).active
        return sheet.max_column

    @dispatch(str, str)
    def get_col_count(self, file_path, sheet_name):
        sheet = openpyxl.load_workbook(file_path).get_sheet_by_name(sheet_name)
        return sheet.max_column

    @dispatch(str, int, int)
    def read(self, file_path, row, col):
        sheet = openpyxl.load_workbook(file_path).active
        cell  = sheet.cell(row, col)
        return cell.value

    @dispatch(str, str, int, int)
    def read(self, file_path, sheet_name, row, col):
        cell = openpyxl.load_workbook(file_path).get_sheet_by_name(sheet_name).cell(row, col)
        return cell.value

    @dispatch(str, int, int, object)
    def write(self, file_path, row, col, data):
        workbook = openpyxl.load_workbook(file_path)
        sheet    = workbook.active
        thin     = Side(border_style='thin', color='00000000')

        sheet.cell(row, col).font   = Font(name='Calibri', size=11, bold=False, color='00000000')
        sheet.cell(row, col).border = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet.cell(row, col).value  = data

        workbook.save(file_path)

    @dispatch(str, int, int, object, str, bool)
    def write(self, file_path, row, col, data, font_color, is_bold):
        workbook = openpyxl.load_workbook(file_path)
        sheet    = workbook.active
        thin     = Side(border_style='thin', color='00000000')

        sheet.cell(row, col).font   = Font(name='Calibri', size=11, bold=is_bold, color=font_color)
        sheet.cell(row, col).border = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet.cell(row, col).value  = data

        workbook.save(file_path)

    @dispatch(str, str, int, int, object)
    def write(self, file_path, sheet_name, row, col, data):
        workbook = openpyxl.load_workbook(file_path)
        sheet    = workbook.get_sheet_by_name(sheet_name)
        thin     = Side(border_style='thin', color='00000000')

        sheet.cell(row, col).font   = Font(name='Calibri', size=11, bold=False, color='00000000')
        sheet.cell(row, col).border = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet.cell(row, col).value  = data

        workbook.save(file_path)

    @dispatch(str, str, int, int, object, str, bool)
    def write(self, file_path, sheet_name, row, col, data, font_color, is_bold):
        workbook = openpyxl.load_workbook(file_path)
        sheet    = workbook.get_sheet_by_name(sheet_name)
        thin     = Side(border_style='thin', color='00000000')

        sheet.cell(row, col).font   = Font(name='Calibri', size=11, bold=is_bold, color=font_color)
        sheet.cell(row, col).border = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet.cell(row, col).value  = data

        workbook.save(file_path)

    def create(self, file_path):
        workbook    = openpyxl.Workbook()
        sheet       = workbook.active
        sheet.title = 'Test Results'
        thin        = Side(border_style='thin', color='00000000')

        sheet['A1'].font                   = Font(name='Calibri', size=12, bold=True, color='00000080')
        sheet['A1'].border                 = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet['A1'].alignment              = Alignment(horizontal='center')
        sheet['A1']                        = 'Scenario'
        sheet.column_dimensions['A'].width = 30

        sheet['B1'].font                   = Font(name='Calibri', size=12, bold=True, color='00000080')
        sheet['B1'].border                 = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet['B1'].alignment              = Alignment(horizontal='center')
        sheet['B1']                        = 'Quote No'
        sheet.column_dimensions['B'].width = 17

        sheet['C1'].font                   = Font(name='Calibri', size=12, bold=True, color='00000080')
        sheet['C1'].border                 = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet['C1'].alignment              = Alignment(horizontal='center')
        sheet['C1']                        = 'Quick Quote Premium'
        sheet.column_dimensions['C'].width = 25

        sheet['D1'].font                   = Font(name='Calibri', size=12, bold=True, color='00000080')
        sheet['D1'].border                 = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet['D1'].alignment              = Alignment(horizontal='center')
        sheet['D1']                        = 'Create Quote Premium'
        sheet.column_dimensions['D'].width = 25

        sheet['E1'].font                   = Font(name='Calibri', size=12, bold=True, color='00000080')
        sheet['E1'].border                 = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet['E1'].alignment              = Alignment(horizontal='center')
        sheet['E1']                        = 'Proposal Download'
        sheet.column_dimensions['E'].width = 23

        sheet['F1'].font                   = Font(name='Calibri', size=12, bold=True, color='00000080')
        sheet['F1'].border                 = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet['F1'].alignment              = Alignment(horizontal='center')
        sheet['F1']                        = 'Schedule Download'
        sheet.column_dimensions['F'].width = 23

        sheet['G1'].font                   = Font(name='Calibri', size=12, bold=True, color='00000080')
        sheet['G1'].border                 = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet['G1'].alignment              = Alignment(horizontal='center')
        sheet['G1']                        = 'Application State'
        sheet.column_dimensions['G'].width = 23

        sheet['H1'].font                   = Font(name='Calibri', size=12, bold=True, color='00000080')
        sheet['H1'].border                 = Border(left=thin, top=thin, right=thin, bottom=thin)
        sheet['H1'].alignment              = Alignment(horizontal='center')
        sheet['H1']                        = 'Status'
        sheet.column_dimensions['H'].width = 12

        workbook.save(file_path)